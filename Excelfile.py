import xlsxwriter as xl;
import openpyxl;
import Capture as cp;

def emptySlot(r, c):
    book = openpyxl.load_workbook('C:/Users/Saleem/Desktop/University/ML/Project/MachineLearning.xlsx')
    sheet = book.active
    entry = False
    while(entry==False):
        a1 = sheet.cell(r, 1)
        b1 = sheet.cell(r,2)
        if(a1.value==None):
            entry = True
            break
        else:
            ws.write_string(r-1,0,a1.value)
            ws.write_string(r-1,1,b1.value)
            r+=1
    return r

wb = xl.Workbook("C:/Users/Saleem/Desktop/University/ML/Project/MachineLearning.xlsx")
ws = wb.add_worksheet("Sheet1")

header = wb.add_format(
    {
     "font_size": 14,
     "font_name": "Arial",
     "bold": True
    }
)

A = ws.write("A1","Registration Number")
ws.set_column("A:A",30)
B = ws.write("B1","Name")
ws.set_column("B:B",30)
r = emptySlot(1,0)
ws.write_string(r-1, 0, cp.face_id)
ws.write_string(r-1, 1, cp.name)

A = ws.write("A1","Registration Number",header)
B = ws.write("B1","Name",header)

        
wb.close()