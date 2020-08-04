import cv2, numpy as np;
import openpyxl;
import time;

start=time.time()
period=8

recognizer = cv2.face_LBPHFaceRecognizer.create();
recognizer.read('C:/Users/Saleem/Desktop/University/ML/Project/trainer/trainer.yml');
face_cas = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
cap = cv2.VideoCapture(0);

while True:
    ret, img = cap.read();
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY);
    faces = face_cas.detectMultiScale(gray, 1.3, 7);
    for (x,y,w,h) in faces:
        cv2.rectangle(img, (x,y), (x+w, y+h), (255,0,0),2);
        roi_gray = gray[y:y + h, x:x + w]
        id=recognizer.predict(roi_gray)
        v = id[0]
        book = openpyxl.load_workbook('C:/Users/Saleem/Desktop/University/ML/Project/MachineLearning.xlsx')
        sheet = book.active
        r=2
        c=1
        a1 = sheet.cell(r, c)
        while(a1.value!=None and a1.value!=v):
            a1 = sheet.cell(r, c)
            if(int(a1.value)==v):
                break
            elif(a1.value==None):
                print("Not Found")
            else:
                r+=1
        #cv2.cv.PutText(cv2.cv.fromarray(img),str(id),(x,y+h),font,(0,0,255));
    cv2.imshow('frame',img);
    if time.time()>start+period:
        break;
    if cv2.waitKey(10) & 0xFF == ord('q'):
        break;

print(sheet.cell(r,c+1).value)
cap.release();
cv2.destroyAllWindows();